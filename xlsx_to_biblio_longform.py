"""
xlsx_to_biblio_longform.py

Converte un Excel (stessa struttura di primiMilleLibri.xlsx ecc.) in un CSV longform:
- una riga per ogni cella valorizzata
- colonne: field, value
- se l'intestazione di colonna è mancante, usa "noName" (anziché A, B, C, ...)

Uso:
  python xlsx_to_biblio_longform.py input.xlsx output.csv

Note:
- Legge tutti i fogli del file.
- Salta le righe completamente vuote.
- Normalizza gli spazi nel valore.
"""

import sys
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


def excel_col_letter(n: int) -> str:
    """1-indexed -> Excel column letters (1->A, 27->AA, ...)."""
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def is_empty_cell(v) -> bool:
    if v is None:
        return True
    if isinstance(v, float) and np.isnan(v):
        return True
    s = str(v).strip()
    return s == "" or s.lower() == "nan"


def xlsx_to_longform(in_xlsx: str, out_csv: str) -> None:
    wb = load_workbook(in_xlsx, read_only=True, data_only=True)

    rows_out = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # Header (prima riga)
        header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        header = [("" if v is None else str(v).strip()) for v in header_cells]

        # Righe dati (da 2 in poi)
        for r in range(2, ws.max_row + 1):
            vals = next(ws.iter_rows(min_row=r, max_row=r, values_only=True))

            # Se la riga è tutta vuota, skip
            if all(is_empty_cell(v) for v in vals):
                continue

            for j, v in enumerate(vals, start=1):
                if is_empty_cell(v):
                    continue

                value = re.sub(r"\s+", " ", str(v).strip())

                raw_header = header[j - 1] if (j - 1) < len(header) else ""
                col_letter = excel_col_letter(j)

                # Se header mancante o “Unnamed…”, marca come noName
                if raw_header == "" or raw_header.lower().startswith("unnamed"):
                    field = "noName"
                else:
                    field = raw_header

                # Ulteriore protezione: se per qualche motivo field è una lettera Excel (A, AA, ...)
                # la convertiamo comunque in noName
                if re.fullmatch(r"[A-Z]{1,3}", field or ""):
                    field = "noName"

                rows_out.append({"field": field, "value": value})

    df = pd.DataFrame(rows_out, columns=["field", "value"])
    df.to_csv(out_csv, index=False, encoding="utf-8-sig")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Uso: python3 xlsx_to_biblio_longform.py input.xlsx output.csv")
        sys.exit(1)

    in_xlsx = sys.argv[1]
    out_csv = sys.argv[2]

    if not Path(in_xlsx).exists():
        raise FileNotFoundError(f"File non trovato: {in_xlsx}")

    xlsx_to_longform(in_xlsx, out_csv)
    print(f"OK: scritto {out_csv}")
